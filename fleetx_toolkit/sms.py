"""SemySMS send helpers and sample-Excel template builders.
Pure logic + I/O only; no Tk. The UI tab wires these into _loop."""
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

from .config import SEMYSMS_API, SEMYSMS_SIM_NAMES


def normalize_number(raw, country_code=""):
    """Trim, strip spaces/dashes and stray Excel '.0'. If country_code is set
    (e.g. '91') and the number has no '+' and no leading country code, prepend
    a '+cc'. Returns '' for blanks."""
    s = str(raw or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(" ", "").replace("-", "")
    if not s:
        return ""
    if country_code and not s.startswith("+"):
        cc = str(country_code).strip().lstrip("+")
        if not s.startswith(cc):
            s = "+" + cc + s
        else:
            s = "+" + s
    return s


def sms_success(resp_json):
    """SemySMS returns {"code":"0","id":...} on success. code may be str or int."""
    if not isinstance(resp_json, dict):
        return False
    return str(resp_json.get("code")) == "0"


def build_sms_params(token, device_id, phone, msg):
    """POST form params for api/3/sms.php."""
    return {"token": token, "device": device_id, "phone": phone, "msg": msg}


# ─────────────── sample Excel templates ───────────────
# Headers MUST match what load_excel_records() produces after it lowercases and
# underscores them, i.e. the keys each _run_* method reads.

_SAMPLES = {
    "sms": {
        "headers": ["Mobile", "Message", "SIM Name"],
        "rows": [
            ["9876543210", "setparam 2004:0", "Airtel 1"],
            ["9811122233", "$TEXT_OP:MAINSERVER:52.223.26.135,4444,2019", "Voda Pulse"],
            ["9800000000", "getgps", ""],   # blank SIM -> tab dropdown default
        ],
        "dropdown_col": 3,                   # C = SIM Name
        "dropdown_values": SEMYSMS_SIM_NAMES,
    },
    "device_add": {
        "headers": ["ID", "SIM", "Device Type", "Serial Number"],
        "rows": [
            ["868020030000001", "8991000012345678901", "FMB920", "SN0001"],
            ["868020030000002", "8991000012345678902", "FMC125", "SN0002"],
        ],
        "dropdown_col": 3,
        "dropdown_values": ["FMB920", "FMC125", "FMB125", "FMB003", "FX10", "FX11"],
    },
    "sim_update": {
        "headers": ["Device ID", "SIM", "Mobile"],
        "rows": [
            ["123456", "8991000012345678901", "9876543210"],
            ["123457", "8991000012345678902", ""],   # Mobile optional -> falls back to SIM
        ],
    },
    "vehicle_map": {
        "headers": ["Device ID", "Vehicle ID"],
        "rows": [
            ["123456", "778001"],
            ["123457", "778002"],
        ],
    },
}


def write_sample(kind, path):
    """Create a formatted sample .xlsx for the given tab kind at `path`."""
    spec = _SAMPLES[kind]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(spec["headers"])
    for c in range(1, len(spec["headers"]) + 1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 24
    for row in spec["rows"]:
        ws.append(row)
    # in-cell dropdown (data validation) if this template defines one
    col = spec.get("dropdown_col")
    if col:
        letter = openpyxl.utils.get_column_letter(col)
        vals = ",".join(spec["dropdown_values"])
        dv = DataValidation(type="list", formula1=f'"{vals}"', allow_blank=True)
        dv.add(f"{letter}2:{letter}1000")
        ws.add_data_validation(dv)
    wb.save(path)
    return path
