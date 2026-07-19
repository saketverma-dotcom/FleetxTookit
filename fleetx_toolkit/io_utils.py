"""Input parsing (pasted IDs, curl), Excel read, and result-log write."""
import datetime
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import LOGS_DIR

XL_GREEN = PatternFill("solid", fgColor="C6EFCE")
XL_RED   = PatternFill("solid", fgColor="FFC7CE")
XL_BOLD  = Font(bold=True)


def parse_pasted_ids(text):
    ids = []
    for chunk in text.replace(",", "\n").split("\n"):
        val = chunk.strip()
        if val and val.replace(".", "").isdigit():
            ids.append(val.split(".")[0])
    return ids

def parse_pasted_pairs(text):
    """Parse lines like 'deviceId vehicleId' or 'deviceId,vehicleId'."""
    pairs = []
    for line in text.strip().split("\n"):
        parts = re.split(r"[,\s\t]+", line.strip())
        nums = [p.split(".")[0] for p in parts if p.strip().replace(".", "").isdigit()]
        if len(nums) >= 2:
            pairs.append((nums[0], nums[1]))
    return pairs

def parse_curl_command(curl_text):
    """Extract commandId, commandName, deviceType, style from a pasted curl."""
    out = {"name": "", "id": "", "deviceType": "FX10", "style": "json"}
    # form style:  --data commandId=xxx   or  --data 'commandId=xxx'
    m = re.search(r"commandId['\"]?\s*[=:]\s*['\"]?([0-9a-fA-F]{20,32})", curl_text)
    if m:
        out["id"] = m.group(1)
    m = re.search(r"commandName['\"]?\s*[=:]\s*['\"]?([\w.\-]+)", curl_text)
    if m:
        out["name"] = m.group(1)
    m = re.search(r"deviceType['\"]?\s*[=:]\s*['\"]?([\w\-]+)", curl_text)
    if m:
        out["deviceType"] = m.group(1)
    # multipart / urlencoded markers → form style
    if "form-data" in curl_text or "--data token=" in curl_text or "urlencoded" in curl_text \
            or "DYNAMIC_COMMAND_SETTING_TRIGGER" in curl_text \
            or "trigger/sendcommands" in curl_text:
        out["style"] = "form"
    return out

def load_excel_column(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
    prefer = ("imei", "id", "vehicle_id", "ticket_id", "device_id", "sim")
    col_idx = next((header.index(c) for c in prefer if c in header), 0)
    vals = []
    for row in rows[1:]:
        v = row[col_idx] if col_idx < len(row) else None
        if v is not None and str(v).strip():
            vals.append(str(v).strip().split(".")[0])
    return vals

def load_excel_records(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:] if any(v for v in row if v is not None)]

def save_result_log(results, columns, label, log_fn):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(columns + ["Status", "Result", "Response", "Timestamp"])
    for cell in ws[1]:
        cell.font = XL_BOLD
    for i in range(1, len(columns) + 5):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions[get_column_letter(len(columns) + 3)].width = 55

    success = fail = 0
    for r in results:
        ok = 200 <= r["status"] < 300
        ws.append(list(r["fields"]) + [r["status"], "SUCCESS" if ok else "FAILED", r["body"], r["ts"]])
        for cell in ws[ws.max_row]:
            cell.fill = XL_GREEN if ok else XL_RED
        if ok: success += 1
        else:  fail += 1

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Operation", "Total", "Success", "Failed", "Run At"])
    for cell in ws2[1]:
        cell.font = XL_BOLD
    ws2.append([label, len(results), success, fail,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    try:
        os.makedirs(LOGS_DIR, exist_ok=True)
    except Exception:
        pass  # FIX-4: handled by the fallback save below
    stamp = f"{datetime.datetime.now():%Y%m%d_%H%M%S}"
    base = f"{label.lower().replace(' ', '_').replace('/', '_')}_log_{stamp}.xlsx"
    fname = os.path.join(LOGS_DIR, base)
    # FIX-4: wb.save() failure (file locked, exe in read-only dir) used to raise
    # inside the worker and lose the entire result set. Fall back to home dir.
    try:
        wb.save(fname)
    except Exception as e:
        fallback = os.path.join(os.path.expanduser("~"), base)
        try:
            wb.save(fallback)
            log_fn(f"  ⚠ Could not write to logs folder ({e}); saved to {fallback}", "err")
            fname = fallback
        except Exception as e2:
            log_fn(f"  ✗ FAILED to save result log anywhere: {e2!r}", "err")
            log_fn(f"  ✓ Success: {success}   ✗ Failed: {fail}\n", "info")
            return
    log_fn(f"\n  Log saved → {fname}", "info")
    log_fn(f"  ✓ Success: {success}   ✗ Failed: {fail}\n", "info")
