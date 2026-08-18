"""Tests for the v3.2 additions: SMS logic, SIM name mapping, sample Excel
templates, shared sensor types. No Tk, no network."""
import openpyxl
import pytest

from fleetx_toolkit import access_control as ac
from fleetx_toolkit import config as cf
from fleetx_toolkit import sms


class TestNormalizeNumber:
    def test_strips_spaces_and_dashes(self):
        assert sms.normalize_number("98765-43 210") == "9876543210"

    def test_strips_excel_float(self):
        assert sms.normalize_number("9876543210.0") == "9876543210"

    def test_blank(self):
        assert sms.normalize_number("") == ""
        assert sms.normalize_number(None) == ""

    def test_country_code_prepended(self):
        assert sms.normalize_number("9876543210", "91") == "+919876543210"

    def test_country_code_not_doubled_when_present(self):
        assert sms.normalize_number("919876543210", "91") == "+919876543210"

    def test_plus_kept(self):
        assert sms.normalize_number("+919876543210", "91") == "+919876543210"

    def test_no_cc_sends_as_is(self):
        assert sms.normalize_number("9876543210") == "9876543210"


class TestSmsSuccess:
    def test_code_zero_string(self):
        assert sms.sms_success({"code": "0", "id": 123}) is True

    def test_code_zero_int(self):
        assert sms.sms_success({"code": 0}) is True

    def test_code_nonzero(self):
        assert sms.sms_success({"code": "1"}) is False

    def test_non_dict(self):
        assert sms.sms_success("boom") is False
        assert sms.sms_success(None) is False


class TestSimMapping:
    def test_all_six_present(self):
        assert set(cf.SEMYSMS_SIMS.values()) == {
            "Airtel 1", "Airtel 2", "Airtel Pulse",
            "Voda Pulse", "Voda Restrict 1", "Voda Restrict 2"}

    def test_name_to_id(self):
        assert cf.sim_id_for_name("Airtel 1") == "355387"
        assert cf.sim_id_for_name("Voda Pulse") == "352969"

    def test_unknown_name(self):
        assert cf.sim_id_for_name("Nokia 3310") == ""
        assert cf.sim_id_for_name("") == ""

    def test_whitespace_tolerant(self):
        assert cf.sim_id_for_name("  Airtel 2  ") == "355386"

    def test_build_params(self):
        p = sms.build_sms_params("tok", "355387", "+91987", "hi")
        assert p == {"token": "tok", "device": "355387",
                     "phone": "+91987", "msg": "hi"}


class TestSampleTemplates:
    @pytest.mark.parametrize("kind,headers", [
        ("sms", ["Mobile", "Message", "SIM Name"]),
        ("device_add", ["ID", "SIM", "Device Type", "Serial Number"]),
        ("sim_update", ["Device ID", "SIM", "Mobile"]),
        ("vehicle_map", ["Device ID", "Vehicle ID"]),
    ])
    def test_headers_match_code_expectations(self, kind, headers, tmp_path):
        p = tmp_path / f"{kind}.xlsx"
        sms.write_sample(kind, str(p))
        wb = openpyxl.load_workbook(p)
        assert [c.value for c in wb.active[1]] == headers

    def test_headers_normalize_to_read_keys(self, tmp_path):
        # a sample must round-trip through load_excel_records to the keys the
        # _run_* methods actually read
        from fleetx_toolkit.io_utils import load_excel_records
        p = tmp_path / "s.xlsx"
        sms.write_sample("device_add", str(p))
        recs = load_excel_records(str(p))
        assert set(recs[0]) >= {"id", "sim", "device_type", "serial_number"}

    def test_sms_sample_has_sim_dropdown(self, tmp_path):
        p = tmp_path / "sms.xlsx"
        sms.write_sample("sms", str(p))
        wb = openpyxl.load_workbook(p)
        dvs = wb.active.data_validations.dataValidation
        assert dvs and any("Airtel 1" in (dv.formula1 or "") for dv in dvs)


class TestSmsRowsExcel:
    """Drives SmsTabMixin._sms_rows without Tk via a minimal stub."""

    class _V:
        def __init__(self, v): self._v = v
        def get(self): return self._v

    def _stub(self, path, cc="", sim="Airtel 1"):
        from fleetx_toolkit.ui.tabs_sms import SmsTabMixin
        s = SmsTabMixin.__new__(SmsTabMixin)
        s.sms_mode = self._V("excel"); s.sms_path = self._V(path)
        s.sms_cc = self._V(cc); s.sms_sim = self._V(sim)
        s.errors = []
        s._load_excel_safe = lambda loader, p: loader(p)
        s._ui_error = lambda t, m: s.errors.append(m)
        return s

    def test_per_row_sim_and_blank_fallback(self, tmp_path):
        p = tmp_path / "s.xlsx"
        sms.write_sample("sms", str(p))
        rows = self._stub(str(p))._sms_rows()
        assert rows[0] == ("9876543210", "setparam 2004:0", "355387")   # Airtel 1
        assert rows[1][2] == "352969"                                   # Voda Pulse
        assert rows[2] == ("9800000000", "getgps", "355387")            # blank -> default

    def test_none_cell_treated_as_blank(self, tmp_path):
        # regression: openpyxl returns None for an empty SIM Name cell;
        # str(None) must not become an "unknown SIM" error.
        p = tmp_path / "n.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Mobile", "Message", "SIM Name"])
        ws.append([9800000000, "getgps", None])
        wb.save(p)
        rows = self._stub(str(p))._sms_rows()
        assert rows == [("9800000000", "getgps", "355387")]

    def test_unknown_sim_name_rejected(self, tmp_path):
        p = tmp_path / "u.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Mobile", "Message", "SIM Name"])
        ws.append(["999", "hi", "Jio Turbo"])
        wb.save(p)
        s = self._stub(str(p))
        assert s._sms_rows() is None
        assert "Jio Turbo" in s.errors[0]

    def test_country_code_applied(self, tmp_path):
        p = tmp_path / "c.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Mobile", "Message", "SIM Name"])
        ws.append(["9876543210", "hi", "Airtel 2"])
        wb.save(p)
        rows = self._stub(str(p), cc="91")._sms_rows()
        assert rows[0][0] == "+919876543210" and rows[0][2] == "355386"

    def test_rows_missing_mobile_or_message_skipped(self, tmp_path):
        p = tmp_path / "m.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Mobile", "Message", "SIM Name"])
        ws.append(["", "hi", "Airtel 1"])       # no mobile
        ws.append(["999", "", "Airtel 1"])      # no message
        ws.append(["888", "ok", "Airtel 1"])    # valid
        wb.save(p)
        rows = self._stub(str(p))._sms_rows()
        assert rows == [("888", "ok", "355387")]


class TestAccessSavePreservesMeta:
    """Regression: saving the access map must NOT wipe _-prefixed meta keys
    (the update pointer + shared sensor list) for the whole team."""

    def test_meta_survives_access_save(self, monkeypatch):
        import json as _json
        from fleetx_toolkit import access_control as acmod

        live = {
            "old@fleetx.io": ["Tickets"],
            "_latest_version": "3.3.1",
            "_download_url": "https://x/e.exe",
            "_sha256": "abc",
            "_sensor_types": ["CUSTOM_A"],
        }
        captured = {}

        class G:
            status_code = 200
            def json(self):
                return {"files": {"fleetx_access.json":
                                  {"content": _json.dumps(live)}}}

        class P:
            status_code = 200

        monkeypatch.setattr(acmod.session, "get",
                            lambda *a, **k: G())
        monkeypatch.setattr(acmod.session, "patch",
                            lambda *a, **k: (captured.update(body=k["json"]) or P()))
        monkeypatch.setattr(acmod, "ACCESS_FILE", "/tmp/_acc_test.json")

        ok, _ = acmod.push_access_to_gist(
            {"new@fleetx.io": ["SMS Command"]}, "tok")
        written = _json.loads(
            captured["body"]["files"]["fleetx_access.json"]["content"])

        assert ok
        assert written["new@fleetx.io"] == ["SMS Command"]
        assert written["_latest_version"] == "3.3.1"
        assert written["_download_url"] == "https://x/e.exe"
        assert written["_sha256"] == "abc"
        assert written["_sensor_types"] == ["CUSTOM_A"]

    def test_caller_meta_keys_ignored_in_favor_of_live(self, monkeypatch):
        # if the caller passes a stale _latest_version, the live Gist wins
        import json as _json
        from fleetx_toolkit import access_control as acmod
        live = {"_latest_version": "9.9.9"}
        captured = {}

        class G:
            status_code = 200
            def json(self):
                return {"files": {"fleetx_access.json":
                                  {"content": _json.dumps(live)}}}

        monkeypatch.setattr(acmod.session, "get", lambda *a, **k: G())
        monkeypatch.setattr(acmod.session, "patch",
                            lambda *a, **k: (captured.update(body=k["json"])
                                             or type("P", (), {"status_code": 200})()))
        monkeypatch.setattr(acmod, "ACCESS_FILE", "/tmp/_acc_test2.json")

        acmod.push_access_to_gist(
            {"u@fleetx.io": [], "_latest_version": "0.0.1"}, "tok")
        written = _json.loads(
            captured["body"]["files"]["fleetx_access.json"]["content"])
        assert written["_latest_version"] == "9.9.9"   # live, not caller's stale


class TestDelayOverride:
    """SMS tab sets a per-run delay override that _current_delay honors."""

    def _fake(self):
        from fleetx_toolkit.ui.app import FleetXToolkit
        class V:
            def __init__(self, v): self.v = v
            def get(self): return self.v
        f = FleetXToolkit.__new__(FleetXToolkit)
        f.delay_var = V("1250")
        f._delay_override = None
        return f

    def test_settings_delay_when_no_override(self):
        f = self._fake()
        assert abs(f._current_delay() - 1.25) < 1e-9

    def test_override_wins(self):
        f = self._fake(); f._delay_override = 5.0
        assert f._current_delay() == 5.0

    def test_zero_delay_allowed_via_override(self):
        f = self._fake(); f._delay_override = 0.0
        assert f._current_delay() == 0.0

    def test_cleared_override_reverts(self):
        f = self._fake(); f._delay_override = 9.0
        assert f._current_delay() == 9.0
        f._delay_override = None
        assert abs(f._current_delay() - 1.25) < 1e-9

    def test_bad_settings_value_falls_back_to_default(self):
        from fleetx_toolkit.config import DELAY_MS
        f = self._fake()
        f.delay_var.v = "abc"
        assert f._current_delay() == DELAY_MS / 1000


class TestSharedSensorTypes:
    def test_empty_when_no_meta(self, monkeypatch):
        monkeypatch.setattr(ac, "_REMOTE_META", {}, raising=False)
        assert ac.get_shared_sensor_types() == []

    def test_reads_list(self, monkeypatch):
        monkeypatch.setattr(ac, "_REMOTE_META",
                            {"_sensor_types": ["CUSTOM_A", "CUSTOM_B"]}, raising=False)
        assert ac.get_shared_sensor_types() == ["CUSTOM_A", "CUSTOM_B"]

    def test_non_list_ignored(self, monkeypatch):
        monkeypatch.setattr(ac, "_REMOTE_META",
                            {"_sensor_types": "oops"}, raising=False)
        assert ac.get_shared_sensor_types() == []


class TestSmsAuth:
    def _seed(self):
        from fleetx_toolkit import sms_auth as A
        store = {"salt": "s4lt", "users": {}}
        return A.apply_user_change(store, "set", "saket.verma@fleetx.io",
                                   password="yvgkqwwpbr", admin=True)

    def test_admin_login(self):
        from fleetx_toolkit import sms_auth as A
        ok, adm, why = A.check_login(self._seed(), "saket.verma@fleetx.io", "yvgkqwwpbr")
        assert ok and adm

    def test_wrong_password(self):
        from fleetx_toolkit import sms_auth as A
        ok, adm, why = A.check_login(self._seed(), "saket.verma@fleetx.io", "nope")
        assert not ok and "Incorrect" in why

    def test_unknown_user(self):
        from fleetx_toolkit import sms_auth as A
        ok, adm, why = A.check_login(self._seed(), "ghost@fleetx.io", "x")
        assert not ok and "Unknown" in why

    def test_email_case_insensitive(self):
        from fleetx_toolkit import sms_auth as A
        ok, _, _ = A.check_login(self._seed(), "  SAKET.VERMA@Fleetx.IO ", "yvgkqwwpbr")
        assert ok

    def test_password_stored_hashed_not_plaintext(self):
        import json
        assert "yvgkqwwpbr" not in json.dumps(self._seed())

    def test_add_reset_delete(self):
        from fleetx_toolkit import sms_auth as A
        s = self._seed()
        s = A.apply_user_change(s, "set", "u@fleetx.io", password="p1", admin=False)
        assert A.check_login(s, "u@fleetx.io", "p1")[0]
        s = A.apply_user_change(s, "set", "u@fleetx.io", password="p2")
        assert not A.check_login(s, "u@fleetx.io", "p1")[0]
        assert A.check_login(s, "u@fleetx.io", "p2")[0]
        s = A.apply_user_change(s, "delete", "u@fleetx.io")
        assert not A.check_login(s, "u@fleetx.io", "p2")[0]

    def test_store_holds_no_token(self):
        # security: the auth Gist must NOT contain a SemySMS token anymore
        import json
        assert "_semysms_token" not in json.dumps(self._seed())
        assert not hasattr(__import__("fleetx_toolkit.sms_auth",
                                      fromlist=["x"]), "get_semysms_token")

    def test_admin_flag_toggle(self):
        from fleetx_toolkit import sms_auth as A
        s = self._seed()
        s = A.apply_user_change(s, "set", "u@fleetx.io", password="p", admin=False)
        assert A.check_login(s, "u@fleetx.io", "p")[1] is False
        s = A.apply_user_change(s, "set", "u@fleetx.io", admin=True)
        assert A.check_login(s, "u@fleetx.io", "p")[1] is True   # pw preserved, admin flipped


class TestPBKDF2AndAutoAuth:
    def test_new_password_uses_pbkdf2(self):
        from fleetx_toolkit import sms_auth as A
        s = A.apply_user_change({"salt": "x", "users": {}}, "set",
                                "u@fleetx.io", password="Secret123")
        rec = s["users"]["u@fleetx.io"]["pw"]
        assert isinstance(rec, dict) and rec["algo"] == "pbkdf2"
        assert rec["rounds"] == A.PBKDF2_ROUNDS
        assert A.check_login(s, "u@fleetx.io", "Secret123")[0]
        assert not A.check_login(s, "u@fleetx.io", "wrong")[0]

    def test_per_user_salts_differ(self):
        from fleetx_toolkit import sms_auth as A
        s = A.apply_user_change({"salt": "x", "users": {}}, "set", "a@f.io", password="same")
        s = A.apply_user_change(s, "set", "b@f.io", password="same")
        assert s["users"]["a@f.io"]["pw"]["salt"] != s["users"]["b@f.io"]["pw"]["salt"]
        assert s["users"]["a@f.io"]["pw"]["hash"] != s["users"]["b@f.io"]["pw"]["hash"]

    def test_legacy_sha256_still_verifies(self):
        from fleetx_toolkit import sms_auth as A
        salt = "c17730ef321706cc"
        legacy = {"salt": salt, "users": {
            "old@fleetx.io": {"pw": A.hash_password("pw123", salt), "admin": True}}}
        ok, adm, _ = A.check_login(legacy, "old@fleetx.io", "pw123")
        assert ok and adm
        assert not A.check_login(legacy, "old@fleetx.io", "nope")[0]

    def test_reset_upgrades_legacy_to_pbkdf2(self):
        from fleetx_toolkit import sms_auth as A
        salt = "s"
        legacy = {"salt": salt, "users": {
            "old@f.io": {"pw": A.hash_password("old", salt), "admin": False}}}
        assert A.is_legacy_hash(legacy["users"]["old@f.io"]["pw"])
        up = A.apply_user_change(legacy, "set", "old@f.io", password="newpw")
        assert not A.is_legacy_hash(up["users"]["old@f.io"]["pw"])
        assert A.check_login(up, "old@f.io", "newpw")[0]

    def test_generate_password(self):
        from fleetx_toolkit import sms_auth as A
        p = A.generate_password()
        assert len(p) == 12
        assert not (set("0O1lI") & set(p))       # no ambiguous chars
        assert A.generate_password() != A.generate_password()

    def test_user_exists_gates_auto_auth(self):
        from fleetx_toolkit import sms_auth as A
        s = A.apply_user_change({"salt": "x", "users": {}}, "set",
                                "vinay@fleetx.io", password="p", admin=False)
        assert A.user_exists(s, "VINAY@Fleetx.IO")        # case-insensitive
        assert not A.user_exists(s, "komal@fleetx.io")    # not on list -> no auto-auth
        assert not A.user_exists({}, "anyone@fleetx.io")

    def test_user_is_admin(self):
        from fleetx_toolkit import sms_auth as A
        s = A.apply_user_change({"salt": "x", "users": {}}, "set",
                                "boss@fleetx.io", password="p", admin=True)
        s = A.apply_user_change(s, "set", "reg@fleetx.io", password="p", admin=False)
        assert A.user_is_admin(s, "boss@fleetx.io")
        assert not A.user_is_admin(s, "reg@fleetx.io")


class TestAutoAuthRouting:
    """v3.11.1: SMS access via FleetX tab grant OR SMS user list; admin only
    from the SMS list / ADMIN_EMAILS."""

    def _store(self):
        from fleetx_toolkit import sms_auth as A
        return A.apply_user_change({"salt": "s", "users": {}}, "set",
                                   "saket.verma@fleetx.io", password="p", admin=True)

    @staticmethod
    def _auto(granted, in_list):
        by_tab = ("SMS Command" in granted) or ("Messaging" in granted)
        return by_tab or in_list

    def test_tab_grant_alone_grants_access(self):
        from fleetx_toolkit import sms_auth as A
        s = self._store()
        assert not A.user_exists(s, "vinay@fleetx.io")
        assert self._auto(["Messaging"], A.user_exists(s, "vinay@fleetx.io"))

    def test_sms_list_alone_grants_access(self):
        from fleetx_toolkit import sms_auth as A
        s = self._store()
        assert self._auto([], A.user_exists(s, "saket.verma@fleetx.io"))

    def test_neither_denies_access(self):
        from fleetx_toolkit import sms_auth as A
        s = self._store()
        assert not self._auto(["Tickets"], A.user_exists(s, "komal@fleetx.io"))

    def test_tab_grant_does_not_confer_sms_admin(self):
        from fleetx_toolkit import sms_auth as A
        s = self._store()
        # granted the tab, but not in the SMS list -> not an SMS admin
        assert not A.user_is_admin(s, "vinay@fleetx.io")

    def test_sms_list_admin_flag_confers_admin(self):
        from fleetx_toolkit import sms_auth as A
        assert A.user_is_admin(self._store(), "saket.verma@fleetx.io")
